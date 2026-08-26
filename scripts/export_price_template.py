"""Write a fill-in-the-prices workbook for the imported Murshid catalogue.

Every row carries the product's External ID, so prices typed here can be
applied back with apply_prices.py without any name matching.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = "import-data/murshid-store/6_prices_TO_FILL.xlsx"
IMD = env["ir.model.data"].sudo()

# One row per BRANCH copy: products are owned by a branch now, so a single row
# per name could only ever price one of the two shops.
rows = []
for p in env["product.template"].sudo().search(
        [("available_in_pos", "=", True)], order="company_id, name"):
    d = IMD.search([("model", "=", "product.template"), ("res_id", "=", p.id)], limit=1)
    rows.append({
        "External ID": d.name if d else "",
        "Branch": p.company_id.name or "ALL BRANCHES",
        "Name": p.name,
        "Product Category": p.categ_id.complete_name,
        "POS Category": ", ".join(p.pos_categ_ids.mapped("name")),
        "Brand": p.brand_id.name or "",
        "Unit": p.uom_id.name or "",
        "Cost": p.standard_price or None,
        "Minimum Selling Price": p.minimum_selling_price or None,
        "Sales Price": p.list_price or None,
        "Maximum Retail Price (MRP)": p.mrp or None,
    })
rows.sort(key=lambda r: (r["Branch"], r["Product Category"], r["Name"]))

HDR = ["Branch", "External ID", "Name", "Product Category", "POS Category", "Brand", "Unit",
       "Cost", "Minimum Selling Price", "Sales Price", "Maximum Retail Price (MRP)"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Prices"
ws.append(HDR)
for c in range(1, len(HDR) + 1):
    ws.cell(row=1, column=c).font = Font(bold=True)
    # Highlight the four columns meant to be typed into.
    if HDR[c - 1] in ("Cost", "Minimum Selling Price", "Sales Price",
                      "Maximum Retail Price (MRP)"):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor="FFF2CC")
for r in rows:
    ws.append([r[h] for h in HDR])

widths = [22, 26, 40, 34, 16, 18, 8, 10, 20, 12, 24]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "D2"
wb.save(OUT)
print(f"wrote {OUT} with {len(rows)} products")
