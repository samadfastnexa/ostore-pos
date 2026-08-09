"""The shipped import templates must work on their own example data.

This exists because they repeatedly did not. Three separate times a template
shipped with rows that failed the moment anyone imported them unchanged:
fractional shortcut quantities on a product sold by the piece, a row named
"Refined Sugar (per Kg)" whose Unit column said Units, and a Packagings cell
using ", " when Odoo's importer does not trim the space. Each was found by
running the file rather than reading it, so that is what these tests do.
"""
import io

from odoo.tests import TransactionCase, tagged

from ..tools import import_templates

MIME_XLSX = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


@tagged('post_install', '-at_install')
class TestImportTemplates(TransactionCase):

    def _load(self, model, content, dryrun=False):
        """Push a workbook through base_import exactly as the web client does."""
        wizard = self.env['base_import.import'].create({
            'res_model': model,
            'file': content,
            'file_name': 'template.xlsx',
            'file_type': MIME_XLSX,
        })
        preview = wizard.parse_preview({
            'headers': True, 'has_headers': True, 'quoting': '"', 'separator': ','})
        self.assertFalse(preview.get('error'), preview.get('error'))
        headers = preview['headers']
        matches = preview.get('matches') or {}
        fields_arg = []
        for index in range(len(headers)):
            match = matches.get(index) or matches.get(str(index)) or []
            fields_arg.append('/'.join(match) or False)
        options = dict(preview['options'],
                       name_create_enabled_fields={}, import_skip_records=[])
        result = wizard.execute_import(fields_arg, headers, options, dryrun=dryrun)
        return headers, fields_arg, result

    # --- every template still maps to the fields we think it does -------

    MODELS = {
        'uom': 'uom.uom',
        'product': 'product.template',
        'product_full': 'product.template',
        'product_category': 'product.category',
        'pos_category': 'pos.category',
        'product_package': 'product.uom',
    }

    def _headers_of(self, key):
        import openpyxl
        content = import_templates.build_workbook(self.env, key)
        sheet = openpyxl.load_workbook(io.BytesIO(content))['Products']
        return [c.value for c in sheet[1] if c.value]

    def test_every_template_header_maps_to_the_intended_field(self):
        """Every shipped header must resolve, and resolve to the RIGHT field.

        Driven off a synthetic one-row sheet rather than the shipped rows: the
        package template cannot carry example rows at all (its Product is
        required and must already exist), and a header-only sheet makes
        parse_preview raise "list index out of range" because it reads a row to
        preview and finds none.
        """
        import xlsxwriter
        for key, model in self.MODELS.items():
            with self.subTest(template=key):
                headers = self._headers_of(key)
                expected = {label: field for label, field, *_r
                            in import_templates.TEMPLATES[key][0]}
                buf = io.BytesIO()
                book = xlsxwriter.Workbook(buf, {'in_memory': True})
                page = book.add_worksheet('Products')
                for col, header in enumerate(headers):
                    page.write(0, col, header)
                    page.write(1, col, 'x')     # any non-empty value: we only
                book.close()                    # care how the header resolves
                wizard = self.env['base_import.import'].create({
                    'res_model': model, 'file': buf.getvalue(),
                    'file_name': 'h.xlsx', 'file_type': MIME_XLSX})
                preview = wizard.parse_preview({
                    'headers': True, 'has_headers': True,
                    'quoting': '"', 'separator': ','})
                self.assertFalse(preview.get('error'), preview.get('error'))
                matches = preview.get('matches') or {}
                for index, header in enumerate(preview['headers']):
                    match = matches.get(index) or matches.get(str(index)) or []
                    self.assertEqual(
                        '/'.join(match), expected.get(header),
                        "%s: header %r resolved to %r" % (key, header, '/'.join(match)))

    def test_shipped_example_rows_import_on_an_empty_database(self):
        """Whatever rows a template ships must work for a brand new shop.

        They did not. The 23-column sheet shipped "Goods / Beverages", "Nestle"
        and "Karachi Beverages Distributors" -- records that existed only
        because the development database had been seeded with them. On a real
        customer's first download that was nine errors before they had typed
        anything. Every example value that needs a pre-existing record is now
        blank, and this test is what keeps it that way.
        """
        for key in ('product', 'product_full', 'uom',
                    'product_category', 'pos_category'):
            with self.subTest(template=key):
                content = import_templates.build_workbook(self.env, key)
                _headers, _fields, result = self._load(
                    self.MODELS[key], content, dryrun=True)
                errors = [m['message'] for m in result.get('messages', [])
                          if m.get('type') == 'error']
                self.assertFalse(errors, "%s: %s" % (key, errors))

    # --- the five-column sheet ------------------------------------------

    def test_product_sheet_shape(self):
        """The columns a shop actually asked for, in a usable order.

        Asserting the exact list rather than a count: this sheet has been cut
        down and built back up several times, and the two failure modes are
        opposite. Too many columns and it is unfillable; too few and the price
        limits, the unit and the supplier all go missing, which is what happened
        when it was trimmed to five.
        """
        import openpyxl
        content = import_templates.build_workbook(self.env, 'product')
        sheet = openpyxl.load_workbook(io.BytesIO(content))['Products']
        self.assertEqual([c.value for c in sheet[1]], [
            'Name', 'Sales Price', 'Cost', 'Barcode', 'Image', 'Unit',
            'Minimum Selling Price', 'Maximum Retail Price (MRP)',
            'Vendors / Vendor', 'Weight', 'External ID',
        ])
        # External ID stays last: it is machinery, not something to fill in.
        self.assertEqual(sheet.cell(1, sheet.max_column).value, 'External ID')
        # Exactly one example row, needing no pre-existing record.
        self.assertEqual(sheet.cell(2, 1).value, 'PVC Pipe 1 inch')
        self.assertIsNone(sheet.cell(3, 1).value, "only one example row expected")

    def test_only_name_and_price_are_required(self):
        required = [label for label, _f, _g, req, *_r
                    in import_templates.TEMPLATES['product'][0] if req]
        self.assertEqual(required, ['Name', 'Sales Price'])

    def test_blank_unit_cell_does_not_break_the_import(self):
        """An empty Unit column means Units, not a crash.

        uom_id is required at database level. Odoo's importer writes False for
        an empty cell in a column that is PRESENT, and False raises
        NotNullViolation before the field default can apply, so a single blank
        Unit cell used to abort the whole file. That is why the column had been
        left out of the sheet altogether.
        """
        product = self.env['product.template'].create({
            'name': 'Blank Unit Product', 'list_price': 100.0, 'uom_id': False})
        self.assertEqual(product.uom_id, self.env.ref('uom.product_uom_unit'))
        # And on write, where the same False arrives from a re-upload.
        product.write({'uom_id': False})
        self.assertTrue(product.uom_id)

    def test_simple_sheet_defaults_fill_the_omitted_columns(self):
        """Omitting Unit and Category must not leave products broken.

        uom_id is required=True; it is safe to omit only because it carries a
        default. categ_id has no default in core at all, which is why this
        module supplies one -- without it every imported row is uncategorised.
        """
        product = self.env['product.template'].create({
            'name': 'Test Pipe', 'list_price': 200.0})
        self.assertTrue(product.uom_id, "uom_id must default, not be left empty")
        self.assertEqual(product.uom_id, self.env.ref('uom.product_uom_unit'))
        self.assertTrue(product.categ_id, "categ_id must default: core supplies none")

    def test_reupload_updates_instead_of_duplicating(self):
        """The External ID column is what makes a mistake recoverable.

        Without it, correcting a price and re-sending the file adds the whole
        catalogue a second time.
        """
        Product = self.env['product.template']
        header = ['Name', 'Sales Price', 'Cost', 'Barcode', 'External ID']

        def sheet(rows):
            import xlsxwriter
            buf = io.BytesIO()
            book = xlsxwriter.Workbook(buf, {'in_memory': True})
            page = book.add_worksheet('Products')
            for col, value in enumerate(header):
                page.write(0, col, value)
            for row, values in enumerate(rows, start=1):
                for col, value in enumerate(values):
                    page.write(row, col, value)
            book.close()
            return buf.getvalue()

        self._load('product.template', sheet([
            ['IT Pipe 1 inch', 200, 150, '', 'it_pipe_1_inch'],
            ['IT Brass Tap', 850, 600, '', 'it_brass_tap'],
        ]))
        after_first = Product.search_count([])
        self.assertEqual(Product.search([('name', '=', 'IT Pipe 1 inch')]).list_price, 200)

        # Same file, one price corrected.
        self._load('product.template', sheet([
            ['IT Pipe 1 inch', 225, 150, '', 'it_pipe_1_inch'],
        ]))
        self.assertEqual(Product.search_count([]), after_first,
                         "re-upload created a duplicate instead of updating")
        self.assertEqual(Product.search([('name', '=', 'IT Pipe 1 inch')]).list_price, 225)

    # --- the layman-facing surfaces --------------------------------------

    def test_valid_values_sheet_uses_labels_not_field_names(self):
        """That sheet exists to be read by a shopkeeper.

        It used to be headed categ_id, pos_categ_ids, seller_ids/partner_id.
        """
        import openpyxl
        content = import_templates.build_workbook(self.env, 'product_full')
        book = openpyxl.load_workbook(io.BytesIO(content))
        self.assertIn(import_templates.VALUES_SHEET, book.sheetnames)
        headings = [c.value for c in book[import_templates.VALUES_SHEET][2] if c.value]
        self.assertTrue(headings, "no pick-lists were written")
        for heading in headings:
            self.assertNotIn('_id', heading, "technical field name leaked: %s" % heading)
            self.assertNotIn('/', heading, "technical field path leaked: %s" % heading)
